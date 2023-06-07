from django.shortcuts import render, redirect
from library.models import Book
from library.forms import BookForm

# Create your views here.

def home_view(request, id=None):
    book_obj = None
    if id:
        book_obj = Book.objects.get(id=id)
    if request.method == 'POST':
        form = BookForm(request.POST, instance=book_obj) #By passing request.POST as the first argument and instance=book as the second argument, you are instructing the form to bind the form data from the POST request and use the data from the book instance to populate the form fields.
        # The form will use the submitted data (request.POST) to update the form fields with the user's input or selected values.
        # The form will also prepopulate the fields with the existing data from the book instance, allowing you to display the current values while editing the book.
        # This approach is commonly used when you want to handle form submission for updating existing data. The form fields are populated with the submitted data and any changes made through the form will be applied to the existing book instance when the form is saved.
        if form.is_valid():
            form.save()
            return redirect('home')
    form = BookForm(instance=book_obj)
    context = {"form": form}
    return render(request, 'home.html', context=context)

def active_books(request):
    all_books = Book.objects.filter(is_active=True).order_by('-id')
    return render(request, 'active_books.html', {"all_books": all_books, "active": True})

def inactive_books(request):
    all_books = Book.objects.filter(is_active=False).order_by('-id')
    return render(request, 'active_books.html', {"all_books": all_books, "inactive": True})

# def update_book(request, id): --- id of the updated book changes
#     book_obj = Book.objects.get(id=id)
#     if request.method == 'POST':
#         form = BookForm(request.POST)
#         if form.is_valid():
#             form.save()
#             book_obj.delete() -- this is to delete the prev obj because this view creates another book objevt - even if we delete it id changes
#             return redirect('active_books')
#     form = BookForm(instance=book_obj) #By passing the instance=book_obj argument, you are instructing the form to use the data from the book_obj instance to populate the form fields. This prepopulates the form with the existing data of the book, allowing you to display the current values and make updates if needed.
#     return render(request, "update.html", context={"form": form})

def hard_delete(request, id):
    book_obj = Book.objects.get(id=id)
    book_obj.delete()
    return redirect('active_books')

def soft_delete(request, id):
    book_obj = Book.objects.get(id=id)
    book_obj.is_active = False
    book_obj.save()
    return redirect('inactive_books')

def restore_book(request, id):
    book_obj = Book.objects.get(id=id)
    book_obj.is_active = True
    book_obj.save()
    return redirect('inactive_books')

from django.http import HttpResponse
import csv
def export_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="all_books.csv"'

    w = csv.writer(response)
    w.writerow(["Name", "Quantity", "Price", "Authour", "Is_Published", "Is_Active"])
    book_obj = Book.objects.all()
    for book in book_obj:
        w.writerow([book.name, book.qty, book.price, book.author, book.is_published, book.is_active])
    books = Book.objects.all().values_list('name','qty', 'price', 'author', 'is_published', 'is_active')
    for book in books:
        w.writerow(book)
    return response


# from django.core.files.storage import FileSystemStorage
# def upload_csv(request):
#     if request.method == 'POST':
#         uploaded_file = request.FILES["document"]
#         file_store = FileSystemStorage()
#         name = file_store.save(uploaded_file.name, uploaded_file) #-- this step is to save the file on server
#         # context = {"url": file_store.url(name)}
#         decoded_file = uploaded_file.read().decode('utf-8').splitlines()
#         expected_header = ["name", "qty", "price", "author", "is_published"]
#         expected_header.sort()
#         actual_header = decoded_file[0].split(",")
#         actual_header.sort()
#         if actual_header != expected_header:
#             return HttpResponse("Headers are not same... please reupload Valid file")
        
#         reader = csv.DictReader(uploaded_file)
#         lst = []
#         for element in reader:
#             if Book.objects.get(name=element.get("name")): # -- validation for unique book name
#                 return HttpResponse(f"Book with name {element.get('name')} already exists...please reupload Valid file")
            
#             is_pub = element.get("is_published")
#             if is_pub == "TRUE":
#                 is_pub = True
#             else:
#                 is_pub = False
#             lst.append(Book(name= element.get("name"), qty= element.get("qty"), price= element.get("price"),author= element.get("author") , is_published=is_pub))            

#         Book.objects.bulk_create(lst)
#         return redirect("active_books")

#     return render(request, "upload.html")

from django.core.files.storage import FileSystemStorage
import chardet # to Detect the encoding of the uploaded file as i was getting error in csv reading

def upload_csv(request):
    if request.method == 'POST':
        uploaded_file = request.FILES.get("document")
        # print(uploaded_file)
        if not uploaded_file:
            return HttpResponse("No file was uploaded")

        file_store = FileSystemStorage()
        name = file_store.save(uploaded_file.name, uploaded_file)
        # print(dir(name), name)
        path = file_store.path(name)
        # Detect the encoding of the uploaded file 
        # print(dir(uploaded_file))
        file = open(path, "r")
        raw_data = file.read()
        # print(raw_data)
        # detection = chardet.detect(raw_data)
        # encoding = detection.get("encoding", "utf-8")  # Set default encoding to 'utf-8'
        # print("encoding", encoding)
        decoded_file = raw_data.splitlines()

        if not decoded_file:  # Check if the decoded_file list is empty
            return HttpResponse("Uploaded file is empty")

        expected_header = ["name", "qty", "price", "author", "is_published"]
        expected_header.sort()

        actual_header = decoded_file[0].split(",")
        actual_header.sort()

        if actual_header != expected_header:
            return HttpResponse("Headers are not the same... please upload a valid file")

        reader = csv.DictReader(decoded_file)

        lst = []
        for element in reader:
            if Book.objects.filter(name=element.get("name")).exists():  # Use filter instead of get to check for existence
                return HttpResponse(f"Book with name {element.get('name')} already exists... please upload a valid file")

            is_pub = element.get("is_published")
            if is_pub == "TRUE":
                is_pub = True
            else:
                is_pub = False

            lst.append(Book(
                name=element.get("name"),
                qty=element.get("qty"),
                price=element.get("price"),
                author=element.get("author"),
                is_published=is_pub
            ))
        # print(lst)
        Book.objects.bulk_create(lst)
        file.close()
        return redirect("active_books")

    return render(request, "upload.html")


from openpyxl import Workbook

def export_active(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="active_books.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    # Sheet header, first row
    columns = ["name", "qty", "price", "author", "is_published" ]

    row = 1
    for i in range(1, len(columns)+1):
        ws.cell(row=row, column=i).value = columns[i-1]

    # Sheet body, remaining rows

    all_data = Book.objects.filter(is_active = True).values_list("name", "qty", "price", "author", "is_published")
    for data in all_data:
        row += 1
        for i in range(1, len(data)+1):
            ws.cell(row=row, column=i).value = data[i-1]

    wb.save(response)
    return response

def export_inactive(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="inactive_books.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    # Sheet header, first row
    columns = ["name", "qty", "price", "author", "is_published" ]

    row = 1
    for i in range(1, len(columns)+1):
        ws.cell(row=row, column=i).value = columns[i-1]

    # Sheet body, remaining rows

    all_data = Book.objects.filter(is_active = False).values_list("name", "qty", "price", "author", "is_published")
    for data in all_data:
        row += 1
        for i in range(1, len(data)+1):
            ws.cell(row=row, column=i).value = data[i-1]

    wb.save(response)
    return response


def upload_text(request):
    if request.method == 'POST':
        uploaded_file = request.FILES.get("document")
        file_store = FileSystemStorage()
        name = file_store.save(uploaded_file.name, uploaded_file)
        path = file_store.path(name)
        file = open(path, "r")
        raw_data = file.readlines()
        context = {"data": raw_data}
        return render(request, "text.html", context=context)

    return render(request, "upload.html")